from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from fpdf import FPDF
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Alignment
from openpyxl.styles import Font

import os
import string
import ssl
import smtplib
import config
import time
import unicodedata
import re
import urllib.request
import sys
import pandas as pd
import pymysql
import openpyxl
import threading
# import subprocess


#############################
###### Fungsi database ######
#############################


# Koneksi ke database
def dbConnectSiap():
    db = pymysql.connect(config.db_host_siap,
                         config.db_username_siap,
                         config.db_password_siap,
                         config.db_name_siap)
    return db


# Mengambil data email per dosen
def getEmailDosen(matkul, kelas, tahun):
    db = dbConnectSiap()
    sql = """
            select d.Email,j.JadwalID,j.TahunID,j.NamaKelas,
            CASE
            WHEN j.ProdiID ='.13.' 
            THEN 'D3 Teknik Informatika'
            WHEN j.ProdiID ='.14.' 
            THEN 'D4 Teknik Informatika'
            WHEN j.ProdiID ='.23.' 
            THEN 'D3 Manajemen Informatika'
            WHEN j.ProdiID ='.33.' 
            THEN 'D3 Akuntansi'
            WHEN j.ProdiID ='.34.' 
            THEN 'D4 Akuntansi Keuangan'
            WHEN j.ProdiID ='.43.' 
            THEN 'D3 Manajemen Pemasaran'
            WHEN j.ProdiID ='.44.' 
            THEN 'D4 Manajemen Perusahaan'
            WHEN j.ProdiID ='.53.' 
            THEN 'D3 Logistik Bisnis'
            WHEN j.ProdiID ='.54.' 
            THEN 'D4 Logistik Bisnis'
            END AS namaprodi,
            j.MKKode,j.DosenID,d.Nama,j.Nama,j.HariID,
            j.JamMulai,j.JamSelesai,j.DosenID 
            from simak_trn_jadwal as j join simak_mst_dosen as d
            where  j.dosenid=d.login and j.MKKode = '""" + matkul + """' 
            and j.NamaKelas = '""" + kelas + """' 
            and TahunID = '""" + tahun + """';
        """
    with db:
        cur = db.cursor()
        cur.execute(sql)
        rows = cur.fetchone()
        if "/" in rows[0]:
            return rows[0][0:rows[0].find("/") - 1].replace(" ", "")
        elif rows[0] != None and rows[0] != "":
            return rows[0].replace(" ", "")
        else:
            return 'NULL'


# Mengambil data matkul per dosen
def getMatkulDosen(dosen, tahun):
    db = dbConnectSiap()
    sql = """
        select j.JadwalID,j.TahunID,j.NamaKelas,
        CASE
        WHEN j.ProdiID ='.13.' 
        THEN 'D3 Teknik Informatika'
        WHEN j.ProdiID ='.14.' 
        THEN 'D4 Teknik Informatika'
        WHEN j.ProdiID ='.23.' 
        THEN 'D3 Manajemen Informatika'
        WHEN j.ProdiID ='.33.' 
        THEN 'D3 Akuntansi'
        WHEN j.ProdiID ='.34.' 
        THEN 'D4 Akuntansi Keuangan'
        WHEN j.ProdiID ='.43.' 
        THEN 'D3 Manajemen Pemasaran'
        WHEN j.ProdiID ='.44.' 
        THEN 'D4 Manajemen Perusahaan'
        WHEN j.ProdiID ='.53.' 
        THEN 'D3 Logistik Bisnis'
        WHEN j.ProdiID ='.54.' 
        THEN 'D4 Logistik Bisnis'
        END AS namaprodi,
        j.MKKode,j.Nama,d.Nama,j.HariID,
        j.JamMulai,j.JamSelesai,j.DosenID 
        from simak_trn_jadwal as j join simak_mst_dosen as d
        where  j.dosenid=d.login and j.DosenID = '"""+dosen+"""' 
        and TahunID = '"""+tahun+"""';
    """
    with db:
        matkul = []

        cur = db.cursor()
        cur.execute(sql)
        rows = cur.fetchall()

        for row in rows:
            matkul.append([row[2], row[3], row[4], row[5], row[6]])

    return pd.DataFrame(matkul, columns=['kelas', 'prodi',
                                         'matkul', 'nama_matkul',
                                         'nama_dosen'])


####################################
###### Membuat file per prodi ######
####################################


def makeFile(driver, list_prodi_ujian, filters):
    launchJadwalUjianMenu(driver)
    getFile(driver, list_prodi_ujian, 
            filters)
    driver.quit()


def getFile(driver, list_prodi_ujian, filters):
    chooseUjian(driver, filters)
    prodi_ujian = Select(
        driver.find_element_by_xpath("//select[@name='prodi']"))

    for prodi_selected in list_prodi_ujian:
        
        checkDirThread(prodi_selected)
        
        if getProdiFromDropdown(driver, prodi_selected, filters):
            printAbsensiUjian(driver, filters, 
                              prodi_selected)


def printAbsensiUjian(driver, filters, prodi):
    time.sleep(3)
    tabel_select = driver.find_element_by_xpath(
        "//table[@cellpadding='4' and @cellspacing='1']/tbody")
    time.sleep(3)
    index = 1
    
    while True:
        try:
            index += 1
            matkul_select = tabel_select.find_element_by_xpath(
                "//tr[" + str(index) + "]/td[8]").text
            time.sleep(1)
            matkul_select = matkul_select.replace(" ", "_")
            matkul_select = matkul_select.replace("-", "_")
            matkul_select = matkul_select.replace("(", "")
            matkul_select = matkul_select.replace(")", "")
            
            kelas_select = tabel_select.find_element_by_xpath(
                "//tr[" + str(index) + "]/td[9]").text
            time.sleep(1)
            
            kode_matkul_select = tabel_select.find_element_by_xpath(
                "//tr[" + str(index) + "]/td[7]").text
            time.sleep(2)
            
            email_select = getEmailDosen(
                kode_matkul_select, kelas_select, filters['tahun'])
            
            kelas_select = int(kelas_select.strip("0"))
            kelas_select = convertKelasByNumber(kelas_select)
            
            filename = ''
            
            if email_select != None:
                filename = "{}-{}-{}-{}-{}-{}".format(filters['tahun'], convertUjianByNumber(
                    filters['jenis']), filters['program'], matkul_select, kelas_select, email_select)
            else:
                filename = "{}-{}-{}-{}-{}-NULL".format(filters['tahun'], convertUjianByNumber(
                    filters['jenis']), filters['program'], matkul_select, kelas_select)

            if os.path.exists('absensi/' + prodi + '/' + filename+config.file_type):
                removeFileThread('absensi/' + prodi + '/' +
                                 filename + config.file_type)
                
            try:
                edit_select = tabel_select.find_element_by_xpath(
                    "//tr[" + str(index) + "]/td[16]/a")
                time.sleep(1)
                
                edit_select.send_keys(Keys.ENTER)
                time.sleep(2)
                
                driver.switch_to.window(driver.window_handles[1])
                time.sleep(2)
                
                url_select = driver.find_element_by_link_text(
                    "Cetak Laporan").get_attribute('href')
                file = 'absensi/' + prodi + '/' + filename + '.txt'
                urllib.request.urlretrieve(
                    url_select, file)
                time.sleep(2)
                
                makePDFOfAbsensiUjian(filename, prodi)
                time.sleep(2)
                
                file = 'absensi/' + prodi + '/' + filename + '.txt'
                if os.path.exists(file):
                    removeFileThread(file)
                msg = 'File ' + filename + config.file_type + ' berhasil dibuat'
                print(msg)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(2)
                
            except NoSuchElementException:
                continue
        except NoSuchElementException:
            break


####################################
###### Membuat file per dosen ######
####################################


def makeFileForDosen(driver, dosens, filters):
    launchJadwalUjianMenu(driver)
    prodis = []
    nama_dosens = []
    
    for dosen in dosens:
        matkul = getMatkulDosen(dosen, filters['tahun'])
        prodis.extend(matkul['prodi'].tolist())
        nama_dosens.extend(matkul['nama_dosen'].tolist())

    prodis = list(set(prodis))
    nama_dosens = list(set(nama_dosens))
    all_ujian = getAllUjian(driver, prodis, 
                            nama_dosens, filters)
    # print(all_ujian)
    
    checkDirsThread(prodis)
    
    msg = "Selesai mengambil semua ujian"
    print(msg)
    
    for dosen in dosens:
        matkul = getMatkulDosen(dosen, filters['tahun'])
        matkul.sort_values(by=['prodi'], inplace=True)
        
        filters.update({"dosen": dosen})
        
        getFileForDosen(driver, all_ujian, matkul, filters)
    driver.quit()


def getFileForDosen(driver, all_ujian, matkul, filters):
    total = len(matkul)
    gagal = 0
    berhasil = 0
    
    for index, row in matkul.iterrows():
        ujian = all_ujian.loc[(all_ujian['matkul'] == row['nama_matkul']) 
                              & (all_ujian['prodi'] == row['prodi'])
                              & (all_ujian['kelas'] == int(row['kelas'].strip("0")))]
        
        if(ujian.empty):
            print("Jadwal "+row['nama_matkul'] +
                  " kelas " + row['kelas'] +
                  " tidak ada di SIAP")
            gagal += 1
        else:
            index_ujian = ujian.iloc[0]['index']
            prodi_ujian = ujian.iloc[0]['prodi']
            
            matkul = {
                "prodi": prodi_ujian,
                "index": str(index_ujian),
                "kelas": ujian.iloc[0]['kelas'],
                "matkul": ujian.iloc[0]['matkul']
            }
            printAbsensiUjianForDosen(driver, matkul, filters)


def printAbsensiUjianForDosen(driver, matkul, filters):
    chooseUjian(driver, filters)
    prodi_ujian = Select(
        driver.find_element_by_xpath("//select[@name='prodi']"))
    
    for prodis in prodi_ujian.options:
        prodi = prodis.text[5:]
        if matkul['prodi'] == prodi:
            prodi_ujian.select_by_value(prodis.text[:2])
            break

    tampil_ujian = driver.find_element_by_xpath("//input[@name='Tampilkan']")
    tampil_ujian.send_keys(Keys.ENTER)
    time.sleep(2)
    
    tabel_select = driver.find_element_by_xpath(
        "//table[@cellpadding='4' and @cellspacing='1']/tbody")
    time.sleep(2)
    
    try:
        dhu_select = tabel_select.find_element_by_xpath(
            "//tr[" + matkul['index'] + "]/td[16]/a")
        print('Jadwal {} kelas {} diproses'.format(
            matkul['matkul'].title(), convertKelasByNumber(matkul['kelas'])))
        time.sleep(2)
        
        matkul_select = tabel_select.find_element_by_xpath(
            "//tr[" + matkul['index'] + "]/td[8]").text
        time.sleep(1)
        
        kelas_select = tabel_select.find_element_by_xpath(
            "//tr[" + matkul['index'] + "]/td[9]").text
        time.sleep(1)
        
        kode_matkul_select = tabel_select.find_element_by_xpath(
            "//tr[" + matkul['index'] + "]/td[7]").text

        email_select = getEmailDosen(
            kode_matkul_select, kelas_select, filters['tahun'])
        
        kelas_select = int(kelas_select.strip("0"))
        kelas_select = convertKelasByNumber(kelas_select)
        matkul_select = matkul_select.replace(" ", "_").replace("-", "_")
        matkul_select = matkul_select.replace("(", "").replace(")", "")
        
        filename = ''
        # Test with your email
        email_select = 'divakrishnam@yahoo.com'
        
        if email_select != None:
            filename = "{}-{}-{}-{}-{}-{}".format(filters['tahun'],
                                                  convertUjianByNumber(
                                                      filters['jenis']),
                                                  filters['program'],
                                                  matkul_select,
                                                  kelas_select,
                                                  email_select)
        else:
            filename = "{}-{}-{}-{}-{}-NULL".format(filters['tahun'],
                                                    convertUjianByNumber(
                                                        filters['jenis']),
                                                    filters['program'],
                                                    matkul_select,
                                                    kelas_select)

        file = 'absensi/'+prodi+'/'+filename+config.file_type
        if os.path.exists(file):
            removeFileThread(file)

        dhu_select.send_keys(Keys.ENTER)
        time.sleep(2)
        
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(2)
        
        url_select = driver.find_element_by_link_text(
            "Cetak Laporan").get_attribute('href')
        file = 'absensi/'+prodi+'/'+filename+'.txt'
        urllib.request.urlretrieve(
            url_select, file)
        time.sleep(2)

        if config.file_type == '.xlsx':
            makeExcelOfAbsensiUjian(filename, prodi)
        elif config.file_type == '.pdf':
            makePDFOfAbsensiUjian(filename, prodi)

        file = 'absensi/'+prodi+'/'+filename+'.txt'
        if os.path.exists(file):
            removeFileThread(file)
        driver.close()
        
        driver.switch_to.window(driver.window_handles[0])
        time.sleep(2)
        
        print('File '+filename+config.file_type+' berhasil dibuat')
        
        if(filters['kirim'] == 'Y'):
            if email_select == 'NULL':
                pass
            else:
                file = {
                    'nama_file': filename+config.file_type,
                    'tujuan': email_select,
                    'matkul': matkul_select,
                    'kelas': kelas_select,
                    'prodi': matkul['prodi'],
                    'jenis': filters['jenis']}
                # print(file)
                sendEmailThread(file)
        #     subprocess.Popen(["python", "kirim_email.py", str(filters)])
    except NoSuchElementException:
        print('Jadwal {} kelas {} belum diatur'.format(
            matkul['matkul'], matkul['kelas']))


##################
###### SIAP ######
##################


# Memilih menu ujian
def launchJadwalUjianMenu(driver):
    driver.get("http://siap.poltekpos.ac.id/")
    user_siap = driver.find_element_by_xpath("//input[@name='user_name']")
    user_siap.send_keys(config.username_siap)
    
    pass_siap = driver.find_element_by_xpath("//input[@name='user_pass']")
    pass_siap.send_keys(config.password_siap)
    
    login_siap = driver.find_element_by_xpath("//input[@name='login']")
    login_siap.send_keys(Keys.ENTER)
    
    jadwal_menu = driver.find_element_by_link_text("Jadwal Ujian 1")
    jadwal_menu.click()


# Memilih ujian
def chooseUjian(driver, filters):
    tahun_ujian = Select(
        driver.find_element_by_xpath("//select[@name='tahun']"))
    tahun_ujian.select_by_value(filters['tahun'])
    
    jenis_ujian = Select(
        driver.find_element_by_xpath("//select[@name='ujian']"))
    jenis_ujian.select_by_value(filters['jenis'])
    
    program_ujian = Select(
        driver.find_element_by_xpath("//select[@name='prid']"))
    program_ujian.select_by_value(filters['program'])

# Memilih prodi


def getProdiFromDropdown(driver, prodi_selected):
    prodi_ujian = Select(
        driver.find_element_by_xpath("//select[@name='prodi']"))
    
    for prodis in prodi_ujian.options:
        prodi = prodis.text[5:].lower()
        prodi_selected = prodi_selected.lower()
        
        if prodi_selected == prodi:
            prodi_ujian.select_by_value(prodis.text[:2])
            return True
            break
        
    print('Prodi {} tidak ada'.format(prodis.text[5:]))
    return False


#############################
###### Dataframe ujian ######
#############################


# Mengambil semua data ujian
def getAllUjian(driver, prodis, dosens, filters):
    chooseUjian(driver, filters)
    list_ujian = pd.DataFrame()
    
    for prodi_selected in prodis:
        if getProdiFromDropdown(driver, prodi_selected):
            tampil_ujian = driver.find_element_by_xpath(
                "//input[@name='Tampilkan']")
            tampil_ujian.send_keys(Keys.ENTER)

            result_generate = genDataFrameUjian(driver, prodi_selected, dosens)
            list_ujian = pd.concat([result_generate, list_ujian])

    return list_ujian


# Membuat dataframe ujian
def genDataFrameUjian(driver, prodi_selected, dosen):
    time.sleep(2)
    tabel_select = driver.find_element_by_xpath(
        "//table[@cellpadding='4' and @cellspacing='1']/tbody")
    time.sleep(2)
    
    index = 1
    dict_data = []
    
    while True:
        try:
            index += 1
            dosen_select = tabel_select.find_element_by_xpath(
                "//tr[" + str(index) + "]/td[13]").text
            time.sleep(1)
            
            if dosen_select in dosen:
                matkul_select = tabel_select.find_element_by_xpath(
                    "//tr[" + str(index) + "]/td[8]").text
                time.sleep(1)
                
                kelas_select = tabel_select.find_element_by_xpath(
                    "//tr[" + str(index) + "]/td[9]").text
                time.sleep(1)
                
                kode_matkul_select = tabel_select.find_element_by_xpath(
                    "//tr[" + str(index) + "]/td[7]").text
                time.sleep(2)
                
                kelas_select = int(kelas_select.strip("0"))
                data = {'prodi': prodi_selected, 
                        'matkul': matkul_select,
                        'kelas': kelas_select, 
                        'kode_matkul': kode_matkul_select, 
                        'index': index}
                dict_data.append(data)
                
            print('.', end='', flush=True)
        except NoSuchElementException:
            print()
            break

    df_data = pd.DataFrame(dict_data)
    return df_data


##########################################
###### Mengirimkan file lewat email ######
##########################################


# Mengirimkan file berdasarkan prodi
def sendFileUjian(list_prodi_ujian, filters):
    for prodi_selected in list_prodi_ujian:
        directory = 'absensi/'+prodi_selected
        for filename in os.listdir(directory):
            if filename.endswith(config.file_type) and filename.startswith(filters['tahun']+'-'+convertUjianByNumber(filters['jenis'])+'-'+filters['program']):
                nama_baru = filename[:-len(config.file_type)].split("-")
                email_dosen = nama_baru[-1]
                if email_dosen == 'NULL':
                    continue
                else:
                    file_baru = nama_baru[0]+'-'+nama_baru[1]+'-'+nama_baru[2]+'-'+nama_baru[3]+'-'+nama_baru[4]+'-'+nama_baru[5]+config.file_type
                    file = {'nama_lama': filename,
                            'prodi': prodi_selected,
                            'nama_baru': file_baru,
                            'tujuan': email_dosen,
                            'ujian': nama_baru[1],
                            'matkul': nama_baru[3],
                            'kelas': nama_baru[4],
                            'prodi': prodi_selected}
                    sendEmailThread(file)

# Mengirimkan file berdasar dosen


def sendFileUjianDosen(dosens, filters):
    for dosen in dosens:
        matkul = getMatkulDosen(dosen, filters['tahun'])
        for prodi_selected in matkul['prodi'].unique():
            directory = 'absensi/'+prodi_selected
            for ind in matkul.index:
                for filename in os.listdir(directory):
                    if filename.endswith(config.file_type) and filename.startswith(filters['tahun']+'-'+convertUjianByNumber(filters['jenis'])+'-'+filters['program']):
                        nama_baru = filename[:-
                                             len(config.file_type)].split("-")
                        
                        email_dosen = nama_baru[-1]
                        
                        matkul_select = matkul['nama_matkul'][ind].replace(
                            " ", "_")
                        matkul_select = matkul_select.replace("-", "_")
                        matkul_select = matkul_select.replace("(", "")
                        matkul_select = matkul_select.replace(")", "")

                        if nama_baru[3] == matkul_select and nama_baru[4] == convertUjianByNumber(int(matkul['kelas'][ind].strip("0"))):

                            if email_dosen == 'NULL':
                                continue
                            else:
                                file_baru = nama_baru[0]+'-'+nama_baru[1]+'-'+nama_baru[2] + \
                                    '-'+nama_baru[3]+'-' + \
                                    nama_baru[4]+config.file_type
                                    
                                file = {'nama_lama': filename,
                                        'prodi': prodi_selected,
                                        'nama_baru': file_baru,
                                        'tujuan': email_dosen,
                                        'ujian': nama_baru[1],
                                        'matkul': nama_baru[3],
                                        'kelas': nama_baru[4],
                                        'prodi': prodi_selected}
                                
                                if sendEmail(file):
                                    print('File '+filename +
                                          ' berhasil dikirim ke '+email_dosen)
                                else:
                                    print('File '+filename +
                                          ' gagal dikirim ke '+email_dosen)


def sendEmailThread(file):
    t = threading.Thread(target=sendEmail,
                         args=(file,))
    t.start()

# Mengirimkan email


def sendEmail(file):
    try:
        subject = "Absensi {} Mata Kuliah {} Kelas {} Prodi {}".format(
            file['jenis'], 
            file['matkul'], 
            file['kelas'], 
            file['prodi'])
        
        body = "Ini file absensi Ujitan UTS Semester Genap oleh iteung ya..., mohon untuk dicek kembali filenya jika ada yang salah mohon untuk diinformasikan ke admin iteung yaa....:) \nAbsensi {} Mata Kuliah {} Kelas {} Prodi {}".format(
            file['jenis'], 
            file['matkul'], 
            file['kelas'], 
            file['prodi'])

        sender_email = config.email_iteung
        
        receiver_email = file['tujuan']
        
        receiver_email = 'divakrishnam@yahoo.com'
        
        password = config.pass_iteung

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Bcc"] = receiver_email

        message.attach(MIMEText(body, "plain"))

        absensifile = file['nama_file']
        path = 'absensi\\'+file['prodi']+'\\'+absensifile
        with open(path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header(
            "Content-Disposition",
            "attachment; filename= %s " % absensifile,
        )

        message.attach(part)

        # Berita acara

        beritafile = 'BERITA_ACARA_UJIAN.docx'
        path = 'absensi\\'+beritafile
        
        with open(path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header(
            "Content-Disposition",
            "attachment; filename= %s " % beritafile,
        )

        message.attach(part)

        text = message.as_string()

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)

        msg = 'File %s berhasil dikirim ke %s' % (absensifile, file['tujuan'])
        print(msg)
        
        path = 'absensi\\'+file['prodi']+'\\'+absensifile
        removeFileThread(path)

    except FileNotFoundError:
        msg = 'File '+absensifile+' tidak ditemukan'
        print(msg)
        pass


################################
###### Membuat file excel ######
################################


# Membuat file excel ujian
def makeExcelOfAbsensiUjian(filename, prodi):
    file = 'absensi/{}/{}.txt'.format(prodi, filename)
    lists = readFile(file)
    wb = openpyxl.Workbook()
    
    sheet = wb.active
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['F'].width = 20
    
    genSubHead(lists['subhead'], sheet)
    genBody(lists['body'], sheet)
    genHead(lists['head'], sheet)
    
    file = 'absensi/{}/{}.xlsx'.format(prodi, filename)
    wb.save(file)


# Membaca file txt ujian
def readFile(filename):
    file = open(filename, 'r+')
    char_re = removeSpecialChar()
    row_list = file.readlines()
    length_row = len(row_list)

    head_list = []
    subhead_list = []
    body_list = []

    index_row = 1
    for row in row_list:
        row = char_re.sub('', row)
        if index_row <= 4 and row != '':
            head_list.append(row.split('\n'))
        elif index_row <= 8 and row != '':
            subhead_list.append(row.split('\n'))
        elif index_row <= length_row-1 and index_row > 8:
            if index_row <= 60:
                body_list.append(row.split('\n'))
            elif index_row >= 72:
                body_list.append(row.split('\n'))
        index_row += 1
    return {'head': head_list,
            'subhead': subhead_list,
            'body': body_list}


# Membuat head file excel ujian
def genHead(head_list, sheet):
    index_head = 0
    index_cell = 2
    for head in head_list:
        sheet.merge_cells('A{}:F{}'.format(index_cell, index_cell))
        value = head_list[index_head][0]
        cell = sheet.cell(row=index_cell, column=1)
        cell.value = value
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=16, bold=True)
        index_cell += 1
        index_head += 1


# Membuat sub-header file excel ujian
def genSubHead(subhead_list, sheet):
    index_subhead = 0
    for subhead in subhead_list:
        for text in subhead:
            if index_subhead == 0:
                text = re.sub(' +', ' ', text)
                text = text.split(':')
                sheet.merge_cells('A5:B5')
                cell = sheet.cell(row=5, column=1)
                cell.value = 'Kode / Mata Kuliah'
                cell.font = Font(bold=True)
                sheet["C5"].value = ': '+text[1][:-8]
                sheet.merge_cells('D5:E5')
                cell = sheet.cell(row=5, column=4)
                cell.value = 'Tanggal'
                cell.font = Font(bold=True)
                sheet["F5"].value = ': '+text[2]
                break
            elif index_subhead == 1:
                text = re.sub(' +', ' ', text)
                text = text.split(':', 2)
                sheet.merge_cells('A6:B6')
                cell = sheet.cell(row=6, column=1)
                cell.value = 'Kelas'
                cell.font = Font(bold=True)
                sheet["C6"].value = ': '+text[1][:-16]
                sheet.merge_cells('D6:E6')
                cell = sheet.cell(row=6, column=4)
                cell.value = 'Jadwal / Ruang'
                cell.font = Font(bold=True)
                sheet["F6"].value = ': '+text[2]
                break
            elif index_subhead == 2:
                text = re.sub(' +', ' ', text)
                text = text.split(':', 2)
                sheet.merge_cells('A7:B7')
                cell = sheet.cell(row=7, column=1)
                cell.value = 'Pengajar'
                cell.font = Font(bold=True)
                sheet["C7"].value = ': '+text[1][:-8]
                sheet.merge_cells('D7:E7')
                cell = sheet.cell(row=7, column=4)
                cell.value = 'Peserta'
                cell.font = Font(bold=True)
                sheet["F7"].value = ': '+text[2]
                break
        index_subhead += 1


# Membuat body file excel ujian
def genBody(body_list, sheet):
    index_cell = 9
    divider = '--'
    for body in body_list:
        for text in body:
            if index_cell == 9:
                cell = sheet["A"+str(index_cell)]
                cell.value = 'No.'
                cell.font = Font(bold=True)
                cell = sheet["B"+str(index_cell)]
                cell.value = 'NPM'
                cell.font = Font(bold=True)
                cell = sheet["C"+str(index_cell)]
                cell.value = 'Nama'
                cell.font = Font(bold=True)
                cell = sheet["D"+str(index_cell)]
                cell.value = 'Hadir'
                cell.font = Font(bold=True)
                cell = sheet["E"+str(index_cell)]
                cell.value = 'Nilai UTS'
                cell.font = Font(bold=True)
                cell = sheet["F"+str(index_cell)]
                cell.value = 'Tanda Tangan'
                cell.font = Font(bold=True)
                index_cell += 1
                break
            elif divider not in text and index_cell > 9:
                text = re.sub(' +', ' ', text)
                if index_cell <= 18:
                    text_list = text.split(' ', 3)
                    del text_list[0]
                else:
                    text_list = text.split(' ', 2)
                cell = sheet["A"+str(index_cell)]
                cell.value = text_list[0]
                cell.alignment = Alignment(horizontal='right')
                sheet["B"+str(index_cell)].value = text_list[1]
                if re.search('[a-zA-Z]', text_list[2][-8:-4]):
                    nama = text_list[2][:-7]
                    kehadiran = text_list[2][-7:-4]
                else:
                    nama = text_list[2][:-8]
                    kehadiran = text_list[2][-8:-4]
                sheet["C"+str(index_cell)].value = nama.strip()
                cell = sheet["D"+str(index_cell)]
                cell.value = kehadiran.strip()
                cell.alignment = Alignment(horizontal='right')
                sheet["E"+str(index_cell)].value = ''
                sheet["F"+str(index_cell)].value = ''
                index_cell += 1
                break

    index_cell += 1
    cell = sheet["C"+str(index_cell)]
    cell.value = 'Tanda Tangan Pengajar :'
    cell.font = Font(bold=True)
    sheet["D"+str(index_cell)].value = ''
    cell = sheet["E"+str(index_cell)]
    cell.value = 'Tanggal :'
    cell.font = Font(bold=True)
    sheet["F"+str(index_cell)].value = ''


##############################
###### Membuat file pdf ######
##############################


# Membuat file pdf
def makePDFOfAbsensiUjian(filename, prodi):
    pdf = setPdfFormat()

    path = 'absensi/{}/{}.txt'.format(prodi, filename)
    file = open(path)
    all_lines = file.readlines()

    length_lines = len(all_lines)

    if length_lines < 61:
        makeCellPdf(pdf, all_lines, 1, length_lines)
    else:
        makeCellPdf(pdf, all_lines, 1, 61)

    if length_lines > 61:
        makeCellPdf(pdf, all_lines, 63, length_lines)
    output = 'absensi/{}/{}.pdf'.format(prodi, filename)
    pdf.output(output)


# Konfigurasi format pdf
def setPdfFormat():
    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_font('Consolas', '',
                 'consola.ttf',
                 uni=True)
    pdf.set_font("Consolas",
                 size=7.4)
    return pdf


# Membuat data perbaris pdf
def makeCellPdf(pdf, all, begin, end):
    char_re = removeSpecialChar()
    pdf.add_page()
    
    for i in range(begin, end):
        pdf.cell(0, 4,
                 txt=char_re.sub('', all[i]), ln=1, border=0)


#############################
###### Fungsi tambahan ######
#############################


# Mengkonversi ujian; misal UTS -> 1
def convertUjianByNumber(ujian):
    ujian = int(ujian)
    if ujian == 1:
        msg = 'UTS'
    elif ujian == 2:
        msg = 'UAS'
    else:
        msg = 'XXX'
    return msg


def convertUjianByLetter(ujian):
    ujian = int(ujian)
    if ujian == 'UTS':
        msg = '01'
    elif ujian == 'UAS':
        msg = '02'
    else:
        msg = '00'
    return msg


# Mengkoversi kelas; misal A -> 1
def convertKelasByNumber(kelas):
    list_kelas = list(string.ascii_lowercase)
    list_nomor = list(range(1, 27))
    dict_kelas = dict(zip(list_nomor, list_kelas))
    for k, v in dict_kelas.items():
        if k == kelas:
            return v.upper()
            break
    msg = "Kelas tidak terdaftar"
    return msg


def convertKelasByLetter(kelas):
    list_kelas = list(string.ascii_lowercase)
    list_nomor = list(range(1, 27))
    dict_kelas = dict(zip(list_nomor, list_kelas))
    for k, v in dict_kelas.items():
        if v == kelas:
            return k
            break
    msg = "Kelas tidak terdaftar"
    return msg


# Mengecek ketersediaan direktori
def checkDirThread(prodi):
    path = 'absensi/{}'.format(prodi)
    if not os.path.exists(path):
        t = threading.Thread(target=os.makedirs,
                             args=(path,))
        t.start()
        t.join()
        msg = 'Direktori {} telah dibuat'.format(prodi)
        print(msg)


def checkDirsThread(prodis):
    threads = []
    
    for prodi in prodis:
        path = 'absensi/{}'.format(prodi)
        if not os.path.exists(path):
            t = threading.Thread(target=os.makedirs,
                                 args=(path,))
            t.start()
            threads.append({'thread': t,
                            'prodi': prodi})

    for t in threads:
        t['thread'].join()
        msg = 'Direktori {} telah dibuat'.format(t['prodi'])
        print(msg)


# Menghapus karakter spesial
def removeSpecialChar():
    all_chars = (chr(i) for i in range(sys.maxunicode))
    control_chars = ''.join(
        c for c in all_chars if unicodedata.category(c) == 'Cc')
    control_char_re = re.compile('[%s]' % re.escape(control_chars))
    return control_char_re


def removeFileThread(namafile):
    t = threading.Thread(target=os.remove,
                         args=(namafile,))
    t.start()
